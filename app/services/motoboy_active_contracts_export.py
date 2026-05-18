"""Exportação Excel de motoboys com contrato vigente (sem data de distrato)."""
from __future__ import annotations

import re
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload

from app.filters import format_currency
from app.models import CONTRACT_TYPE_MOTOBOY, Contract
from app.models.supplier import client_display_label


EXPORT_HEADERS = (
    "Motoboy",
    "Cliente",
    "Data início",
    "Data distrato",
    "Valor prestação (R$)",
    "CPF",
    "CNPJ",
    "Cidade",
    "Estado (UF)",
    "Placa moto",
    "Conta bancária / PIX",
    "Telefone de contato",
)


def _digits_only(value: str | None) -> str:
    return re.sub(r"\D", "", value or "")


def _format_cpf(value: str | None) -> str:
    digits = _digits_only(value)
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return (value or "").strip()


def _format_cnpj(value: str | None) -> str:
    digits = _digits_only(value)
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    return (value or "").strip()


def _format_date(value: date | None) -> str:
    if not value:
        return ""
    return value.strftime("%d/%m/%Y")


def _format_decimal_br(value: Decimal | float | int | None) -> str:
    if value is None:
        return ""
    return format_currency(value)


def _active_motoboy_contracts_query():
    return (
        Contract.query.filter_by(contract_type=CONTRACT_TYPE_MOTOBOY)
        .filter(Contract.end_date.is_(None))
        .options(
            joinedload(Contract.supplier),
            joinedload(Contract.other_supplier),
        )
        .order_by(Contract.supplier_id, Contract.start_date.desc())
    )


def _contract_row(contract: Contract) -> tuple[Any, ...]:
    motoboy = contract.supplier
    client = contract.other_supplier
    return (
        (motoboy.name if motoboy else "").strip(),
        client_display_label(client) if client else "",
        _format_date(contract.start_date),
        _format_date(contract.end_date),
        _format_decimal_br(contract.service_value),
        _format_cpf(motoboy.document if motoboy else None),
        _format_cnpj(motoboy.document_secondary if motoboy else None),
        (motoboy.city if motoboy else "") or "",
        (motoboy.state if motoboy else "") or "",
        (motoboy.bike_plate if motoboy else "") or "",
        (motoboy.bank_account_pix if motoboy else "") or "",
        (motoboy.contact_phone if motoboy else "") or "",
    )


def build_active_motoboy_contracts_xlsx() -> bytes:
    contracts = _active_motoboy_contracts_query().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos ativos"

    header_font = Font(bold=True)
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    for row_idx, contract in enumerate(contracts, start=2):
        for col_idx, value in enumerate(_contract_row(contract), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(EXPORT_HEADERS) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(EXPORT_HEADERS[col_idx - 1])
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            cell_val = row[0].value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[letter].width = min(max_len + 2, 48)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

